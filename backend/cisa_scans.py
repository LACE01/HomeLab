"""CISA Web Scans XLSX importer.

Parses the Qualys WAS export (CISA web-app scan report) and ingests each row as a
finding in our system, using `(vuln_id, web_application, url)` as the canonical
dedup key so re-uploads add only new findings.
"""
import io
import uuid
from datetime import datetime, timezone, timedelta

from openpyxl import load_workbook

from scoring import compute_sla_days


# Map Qualys WAS severity to ours
_SEV_MAP = {
    5: "Critical", "5": "Critical",
    4: "High", "4": "High",
    3: "Medium", "3": "Medium",
    2: "Low", "2": "Low",
    1: "Info", "1": "Info",
    "Critical": "Critical", "High": "High", "Medium": "Medium", "Low": "Low",
}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_sev(raw) -> str:
    if raw is None:
        return "Medium"
    s = _SEV_MAP.get(raw)
    if s:
        return s
    return _SEV_MAP.get(int(raw), "Medium") if isinstance(raw, (int, float)) else "Medium"


async def import_cisa_scans_xlsx(db, content: bytes, source_label: str | None = None) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    required = ("VULN_ID", "NAME", "WEB APPLICATION", "URL", "SEVERITY")
    missing = [c for c in required if c not in col]
    if missing:
        return {"status": "failed", "error": f"Missing columns: {missing}", "imported": 0}

    started_at = _now_iso()
    created = 0
    updated = 0
    errors: list = []
    seen_assets: dict = {}  # web_application → asset_id

    def cell(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        if not row or not cell(row, "VULN_ID"):
            continue
        try:
            vuln_id = str(cell(row, "VULN_ID")).strip()
            name = str(cell(row, "NAME") or "").strip()
            web_app = str(cell(row, "WEB APPLICATION") or "").strip() or "Unknown Web App"
            url = str(cell(row, "URL") or "").strip()
            severity = _norm_sev(cell(row, "SEVERITY"))
            cvss = cell(row, "BASE CVSS")
            cwe = cell(row, "CWE")
            cwe = str(cwe).strip() if cwe else None
            if cwe and not cwe.startswith("CWE-") and cwe.isdigit():
                cwe = f"CWE-{cwe}"
            cve = cell(row, "CVE")
            cve = str(cve).strip() if cve else None
            description = str(cell(row, "DESCRIPTION") or "").strip()
            impact = str(cell(row, "IMPACT") or "").strip()
            solution = str(cell(row, "SOLUTION") or "").strip()
            vuln_type = str(cell(row, "VULN TYPE") or "").strip()
            first_det = cell(row, "FIRST DETECTION")
            last_det = cell(row, "LAST DETECTION")

            # Upsert web-app "asset"
            if web_app not in seen_assets:
                existing_asset = await db.assets.find_one({"hostname": web_app}, {"_id": 0})
                if existing_asset:
                    seen_assets[web_app] = existing_asset["id"]
                else:
                    aid = str(uuid.uuid4())
                    await db.assets.insert_one({
                        "id": aid, "hostname": web_app, "ip": None,
                        "asset_type": "web_app", "environment": "production",
                        "criticality": "high", "exposure": "internet",
                        "platform": "Web", "operating_system": "WebApp",
                        "owner_team": "Unassigned", "tags": ["cisa-web-scan"],
                        "status": "active", "created_at": _now_iso(),
                        "ownership_confidence": 0.4,
                        "ownership_rationale": "Auto-created from CISA Web Scan import",
                    })
                    seen_assets[web_app] = aid
            asset_id = seen_assets[web_app]

            canonical = f"WAS::{vuln_id}::{web_app}::{url}"
            existing = await db.findings.find_one({"canonical_key": canonical}, {"_id": 0})

            base = {
                "source_tool": "CISA Web Scan (Qualys WAS)",
                "source_observation_id": vuln_id,
                "source_native_id": vuln_id,
                "title": name or f"Web vulnerability {vuln_id}",
                "description": description or impact,
                "severity": severity,
                "cve": cve, "cwe": cwe,
                "cvss_score": float(cvss) if isinstance(cvss, (int, float)) else None,
                "epss_score": 0,
                "kev_flag": False, "rti": [],
                "remediation": solution,
                "consequence": impact,
                "business_impact": impact,
                "detection_logic": vuln_type,
                "asset_id": asset_id, "asset_hostname": web_app, "asset_ip": None,
                "asset_criticality": "high", "asset_exposure": "internet",
                "asset_environment": "production", "asset_os": "WebApp",
                "internet_facing": True,
                "owner_team": "Unassigned", "ownership_confidence": 0.4,
                "product_id": None, "product_name": web_app,
                "last_seen_at": str(last_det) if last_det else _now_iso(),
                "last_changed_at": _now_iso(),
                "imported_at": _now_iso(), "detection_channel": "xlsx_upload",
                "url_path": url, "vuln_type": vuln_type,
                "tags": ["cisa-web-scan", "internet_facing"],
            }

            if existing:
                base["status"] = existing["status"]
                base["reopened_count"] = existing.get("reopened_count", 0)
                base["first_seen_at"] = existing.get("first_seen_at") or str(first_det or _now_iso())
                base["canonical_key"] = canonical
                base["risk_score"] = existing.get("risk_score", 50)
                await db.findings.update_one({"id": existing["id"]}, {"$set": base})
                updated += 1
            else:
                sla_d = compute_sla_days(severity, "high")
                first_seen_v = str(first_det) if first_det else _now_iso()
                try:
                    _fs = datetime.fromisoformat(first_seen_v.replace("Z", "+00:00")) if "T" in first_seen_v else datetime.now(timezone.utc)
                except Exception:
                    _fs = datetime.now(timezone.utc)
                new = {
                    "id": str(uuid.uuid4()), "canonical_key": canonical,
                    "first_seen_at": first_seen_v,
                    "reopened_count": 0,
                    "status": "New", "validation_status": "pending",
                    "sla_days": sla_d,
                    "due_at": (_fs + timedelta(days=sla_d)).isoformat(),
                    "compliance_scope": [], "advisory_links": [], "exploit_references": [],
                    "risk_score": {"Critical": 90, "High": 70, "Medium": 50, "Low": 30, "Info": 10}[severity],
                    **base,
                }
                await db.findings.insert_one(new)
                created += 1
        except Exception as e:
            errors.append({"row_vuln_id": str(cell(row, "VULN_ID") or "?"), "error": str(e)})

    # Surface in import_jobs feed
    await db.import_jobs.insert_one({
        "id": str(uuid.uuid4()),
        "source_name": source_label or "CISA Web Scan (manual upload)",
        "mode": "xlsx_upload",
        "status": "success" if not errors or created or updated else "failed",
        "request_id": f"cisa_{uuid.uuid4().hex[:12]}",
        "started_at": started_at, "finished_at": _now_iso(),
        "created_count": created, "updated_count": updated,
        "deduplicated_count": updated, "failed_count": len(errors),
        "retry_count": 0, "errors": errors[:50],
    })

    return {"status": "success", "created": created, "updated": updated,
            "web_apps": len(seen_assets), "errors": errors[:20],
            "started_at": started_at, "finished_at": _now_iso()}
